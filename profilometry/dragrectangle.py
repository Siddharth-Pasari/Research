import math
import tkinter as tk
from tkinter import filedialog

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import pandas as pd
import openpyxl
from openpyxl.styles import Font


title = "TITLE"

def update_difference_list(data_measurements):
    global difference_list

    if len(difference_list[-1]) >= ftnumber:
        difference_list.append([])
        
    difference_list[-1].append(data_measurements[0][3])

def update_excel_with_data(data_measurements, file_path, num):
    print(data_measurements, num)

    df = pd.DataFrame(data_measurements, columns=['Num', 'Top', 'Bottom', 'Difference'])

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        try:
            startrow = writer.sheets['Sheet1'].max_row
        except KeyError:
            startrow = 0

        if (num - 1) % ftnumber == 0 and startrow != 0:  # Assuming ftnumber is defined somewhere else
            startrow += 1

        if num == 1:
            ws = writer.sheets['Sheet1']
            ws.insert_rows(startrow, amount=1)  # Insert a blank row above the title
            ws.cell(row=startrow + 1, column=1, value=title).font = Font(bold=True)  # Write the title and make it bold
            headers = ['Num', 'Top', 'Bottom', 'Difference']
            startrow += 1  # Increment startrow so headers will be below the title
        else:
            headers = False
        
        df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=headers)

def plot_2d_slice(height_values, max_val, excel_path):
    x_values = np.arange(len(height_values))

    fig, ax = plt.subplots()
    ax.plot(x_values, height_values, 'o-')

    ax.set(xlabel='X', ylabel='Height', title='2D Representation of Data Slice')
    ax.grid()

    marker, = ax.plot([0], [height_values[0]], marker='o', color="red", zorder=5)
    cursor_line = ax.axvline(x=0, color='red', linestyle='--', lw=1)

    y_value_annotation = ax.annotate('', xy=(0, 0), xytext=(10, 10),
                                     textcoords="offset points",
                                     bbox=dict(boxstyle="round4", fc="cyan", ec="black", lw=1))

    bottom_val = None

    def on_move(event):
        if not event.inaxes:
            return
        x, y = event.xdata, event.ydata
        nearest_x_index = np.clip(np.searchsorted(x_values, x), 1, len(x_values) - 1)
        x0, x1 = x_values[nearest_x_index - 1], x_values[nearest_x_index]
        y0, y1 = height_values[nearest_x_index - 1], height_values[nearest_x_index]
        distance_to_x0 = abs(x - x0)
        distance_to_x1 = abs(x - x1)
        index = nearest_x_index - 1 if distance_to_x0 < distance_to_x1 else nearest_x_index

        marker.set_data([x_values[index]], [height_values[index]])
        cursor_line.set_xdata([x_values[index]])  # Provide a sequence with a single value

        y_value_annotation.set_text(f'({x_values[index]:.2f}, {height_values[index]:.2f})')
        y_value_annotation.xy = (x_values[index], height_values[index])

        plt.draw()

    def on_click(event):
        global number
        number += 1
        if not event.inaxes:
            return
        
        nonlocal bottom_val
        bottom_val = height_values[np.clip(np.searchsorted(x_values, event.xdata), 1, len(x_values) - 1) - 1]

        difference = max_val - bottom_val

        data_measurements = [(number, max_val, bottom_val, difference)]

        DragRectangle.update_difference_list(DragRectangle, data_measurements)
        update_excel_with_data(data_measurements, excel_path, number)

        plt.close()

    fig.canvas.mpl_connect('motion_notify_event', on_move)
    fig.canvas.mpl_connect('button_press_event', on_click)

    ax.autoscale_view()  # Auto rescale the view after adding the marker
    plt.show()


class DragRectangle:

    difference_list = [[]]

    def __init__(self, ax, x_values, y_values, data, path, num, ftnum, titlet):
        global number, ftnumber, title, difference_list
        self.ax = ax
        self.x_values = x_values
        self.y_values = y_values
        self.data = data
        self.path = path
        self.num = num
        number = num
        self.ftnum = ftnum
        ftnumber = ftnum
        title = titlet
        self.rect = Rectangle((0, 0), 0, 0, linewidth=1, edgecolor='r', facecolor='none')
        self.is_dragging = False
        self.press_event = None
        self.selected_indices = []
        
    def connect(self):
        self.cidpress = self.ax.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.cidrelease = self.ax.figure.canvas.mpl_connect('button_release_event', self.on_release)
        self.cidmotion = self.ax.figure.canvas.mpl_connect('motion_notify_event', self.on_motion)

    def update_difference_list(self, data_measurements):
        if len(self.difference_list[-1]) >= ftnumber:
            self.difference_list.append([])
        self.difference_list[-1].append(data_measurements[0][3])

    def on_press(self, event):
        if event.inaxes != self.ax:
            return

        if event.button == 3:  # Right-click event
            self.on_right_click(event)
            return
        
        if event.button == 2:  # Middle-click event
            self.on_middle_click(event)
            return

        self.press_event = event
        self.rect.set_width(0)
        self.rect.set_height(0)
        self.rect.set_xy((event.xdata, event.ydata))
        self.ax.add_patch(self.rect)
        self.is_dragging = True

    def on_right_click(self, event):
        global number
        number += 1
        data_measurements = [(number, "NoVal", "NoVal", "NoVal")]

        update_excel_with_data(data_measurements, self.path, number)

    def on_middle_click(self, event):
        global number
        number -= 1
        file_path = self.path
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        df = pd.read_excel(file_path)
        last_row_index = df.last_valid_index()

        if last_row_index is not None:
            sheet.delete_rows(last_row_index + 2)

        print("deleted")
        workbook.save(file_path)

    def on_motion(self, event):
        if not self.is_dragging or event.inaxes != self.ax:
            return

        width = event.xdata - self.press_event.xdata
        height = event.ydata - self.press_event.ydata

        self.rect.set_width(width)
        self.rect.set_height(height)
        self.ax.figure.canvas.draw()

    def on_release(self, event):
        if not self.is_dragging or event.inaxes != self.ax:
            return

        self.is_dragging = False
        self.ax.figure.canvas.draw()

        x0, y0 = self.press_event.xdata, self.press_event.ydata
        x1, y1 = event.xdata, event.ydata

        if x0 > x1:
            x0, x1 = x1, x0
        if y0 > y1:
            y0, y1 = y1, y0

        x_indices = [int(i) for i in range(math.ceil(x0), math.floor(x1))]
        y_indices = [int(i) for i in range(math.ceil(y0), math.floor(y1))]

        self.selected_indices = []
        for j in x_indices:
            row_values = [self.data[j, i] for i in y_indices]
            self.selected_indices.append(row_values)
        
        self.findImportantValues()

    def findImportantValues(self):
        if not self.selected_indices:
            return np.nan, np.nan

        max_value = float('-inf')
        max_list = None
        for sublist in self.selected_indices:
            sublist_max = max(sublist)
            if sublist_max > max_value:
                max_value = sublist_max
                max_list = sublist
        
        if max_list is None:
            return np.nan, np.nan
        
        plot_2d_slice(max_list, max_value, self.path)   
