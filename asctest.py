import os
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

os.chdir(r'C:\Users\Siddharth Pasari\OneDrive\Documents\Research\profilometry')
# Load the ASC file and skip rows until "RAW_DATA 3 2400400"
with open(r'C:\Users\Siddharth Pasari\OneDrive\Documents\Research\profilometry\p2.ASC', 'r') as file:
    lines = file.readlines()
    start_index = lines.index("RAW_DATA\t3\t2400400\t\n") + 1

# Process the data and format into a table
data = np.loadtxt(r'C:\Users\Siddharth Pasari\OneDrive\Documents\Research\profilometry\p2.ASC', skiprows=start_index)

# Create a table with a new cell in the row for each space or row in the ASC file
table = []
row = []
for row_values in data:
    for value in row_values:
        if len(row) < 6001:
            row.append(value)
        else:
            table.append(row)
            row = [value]

# Add the last row to the table
if row:
    table.append(row)

# Specify the file name for the XLS file
xls_file_name = 'p2_formatted.xlsx'

# Check if the XLS file already exists
if os.path.exists(xls_file_name):
    os.remove(xls_file_name)  # Delete the existing file

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Determine the min and max values in the dataset
data_min = np.min(data)
data_max = np.max(data)

# Define colors for the gradient fill
color_min = "FF0000"  # Red
color_max = "0000FF"  # Blue

# Define cell size (in points) for square cells
cell_size = 10  # Adjust as needed

# Write the formatted table to the Excel sheet with gradient fill colors and resized cells
for row_idx, row_data in enumerate(table, start=1):
    for col_idx, cell_data in enumerate(row_data, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=cell_data)
        # Calculate the hue based on the cell's value
        hue = int((cell_data - data_min) / (data_max - data_min) * 255)
        # Define the fill color based on the hue
        fill_color = f"{hue:02X}0000" if hue <= 128 else f"FF{255 - hue:02X}00"
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        # Resize the cell to be a square
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = cell_size
        sheet.row_dimensions[row_idx].height = cell_size

# Save the Excel workbook to a file
workbook.save(xls_file_name)

print(f"Formatted table with gradient fill colors and resized cells written to '{xls_file_name}'")
